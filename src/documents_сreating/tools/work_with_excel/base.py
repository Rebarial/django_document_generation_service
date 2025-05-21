from abc import ABC, abstractmethod
import os
from io import BytesIO
from documents_сreating.models.base import BaseModel
from openpyxl.utils import range_boundaries, get_column_letter
from openpyxl.utils.cell import column_index_from_string, coordinate_from_string
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.pagebreak import Break
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
from django.core.files.storage import default_storage
from documents_сreating.models.organization import Organization
from typing import BinaryIO
import math
import subprocess
import locale
from pytils.numeral import rubles

locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')

class BaseExcelDocumentCreate(ABC):
    """
    Базовый класс для создания excel документов
    """

    def __init__(self, document_dict: dict, template_path: str):
        self.document_dict: dict = document_dict
        self.template_path: str = template_path

    def get_nested_attribute(self, document, item):

        try:
            dk = document
            for it in item.split("."):
                dk = getattr(dk, it)
            return dk
        except:
            return None


    def inn_or_kpp(self, text: str) -> str:
        if ("inn_" in text or 
            "_inn" in text):
            return "inn"
        if ("kpp_" in text or
            "_kpp" in text):
            return "kpp"
        return ""

    def fill_doc(self, document: BaseModel, sheet: Worksheet, offset: list, inn_field: str) -> None:
        """
        Заполняет лист excel данными из document сопоставляя по self.document_dict
        """

        offset = offset

        if "Defoult_items" in self.document_dict:
            
            for item_data in self.document_dict["Defoult_items"]:
                
                items_list = list(self.get_nested_attribute(document,item_data["items_model_name"]).all().values())
                if items_list:
                    offset_local = self.add_document_itmes(sheet, list(self.get_nested_attribute(document,item_data["items_model_name"]).all().values()), offset, item_data, document) #Количество строк товаров
                    offset.append({
                        "cell_itmes_number": item_data["cell_itmes_number"],
                        "offset": offset_local
                    })

        if "Images" in self.document_dict and document.is_stamp:

            organization = Organization.objects.filter(inn=self.get_nested_attribute(document, inn_field)).first()

            #Добавление печати и подписи
            if organization:
                for image in self.document_dict["Images"]:
                    if self.get_nested_attribute(organization, image["type"]):
                        self.add_image(sheet, self.get_nested_attribute(organization, image["type"]).name, image["width"], image["height"], self.get_cell_ref(image["cell"], offset))

                #!!Необходимо добавить заполнение КПП если ИНН отсутствует
        if "Concatenation" in self.document_dict:
            for key, cell_ref in self.document_dict["Concatenation"].items():
                if self.get_nested_attribute(document, key):
                    cell = self.get_cell_ref(cell_ref, offset)
                    if sheet[cell].value != None:
                        #sheet[cell].value += f", {self.get_nested_attribute(document, key)}"
                        if sheet.values == "":
                            sheet[cell].value += f"{self.get_nested_attribute(document, key)}"
                        else:
                            sheet[cell].value += f"\n{self.get_nested_attribute(document, key)}"
                        value = sheet[cell].value
                        self.row_height_from_content(sheet, value, cell)
                    else:
                        sheet[cell] = str(self.get_nested_attribute(document, key))
                        value = sheet[cell].value
                        self.row_height_from_content(sheet, value, cell)

        #!!Необходимо добавить заполнение КПП если ИНН отсутствует
        if "Raw_data" in self.document_dict:
            for key, cell_ref in self.document_dict["Raw_data"].items():
                if self.get_nested_attribute(document, key):
                    value = str(self.get_nested_attribute(document, key))
                    inn_kpp = self.inn_or_kpp(value)
                    if inn_kpp == "inn":
                        value = "ИНН " + value
                    elif inn_kpp == "kpp":
                        value = "КПП" + value
                    if isinstance(cell_ref, list):
                        for cell_item in cell_ref:
                            cell = self.get_cell_ref(cell_item, offset)
                            sheet[cell] = value
                            self.row_height_from_content(sheet, value, cell)
                    else:
                        cell = self.get_cell_ref(cell_ref, offset)
                        sheet[cell] = value
                        self.row_height_from_content(sheet, value, cell)

        #!!Изменить склонения месяцов дат
        if "Date" in self.document_dict:
            if "day" in self.document_dict["Date"]:
                for key, cell_ref in self.document_dict["Date"]["day"].items():
                    if self.get_nested_attribute(document, key):
                        sheet[self.get_cell_ref(cell_ref, offset)] = str(self.get_nested_attribute(document, key).day)
            if "month" in self.document_dict["Date"]:
                for key, cell_ref in self.document_dict["Date"]["month"].items():
                    if self.get_nested_attribute(document, key):
                        sheet[self.get_cell_ref(cell_ref, offset)] = str(self.get_nested_attribute(document, key).strftime('%B'))

            if "year" in self.document_dict["Date"]:
                for key, cell_ref in self.document_dict["Date"]["year"].items():
                    if self.get_nested_attribute(document, key):
                        sheet[self.get_cell_ref(cell_ref, offset)] = str(self.get_nested_attribute(document, key).year)

            if "fromat" in self.document_dict["Date"]:
                for key, cell_ref in self.document_dict["Date"]["fromat"].items():
                    if self.get_nested_attribute(document, key):
                        sheet[self.get_cell_ref(cell_ref, offset)] = f'"{self.get_nested_attribute(document, key).day}" {self.get_nested_attribute(document, key).strftime("%B %Y г.")}'


    @abstractmethod
    def create_excel_document(self, document: BaseModel) -> BinaryIO:
        """
        Создает excel документ на основе шаблона и словаря заполнения
        PS сейчас возвращает pdf
        """
        pass
    
    def get_row_with_offset(self, row: int, offset: list) -> int:
        sum = 0
        for item in offset:
            if "cell_itmes_number" in item and row > item["cell_itmes_number"]:
                sum += item["offset"]
        
        return row + sum


    def get_cell_ref(self, key: tuple[str, int], offset: dict) -> str:
        """
        Возвращает строковое представление ссылки excel на основе кортежа, учитывая смещение при добавлении items
        """
        num = key[1]
        return f"{key[0]}{self.get_row_with_offset(num, offset)}"

    def find_nearest_lesser_or_equal(self, numbers: tuple, target: int) -> int:

        lesser_or_equal_numbers = [num for num in numbers if num <= target]
        
        if not lesser_or_equal_numbers:
            return None
        
        return max(lesser_or_equal_numbers)

    def add_rows_break(self, sheet: Worksheet, break_points: list, offset: list) -> None:

        offset_sum = 0
        for item in offset:
            items_row = item["cell_itmes_number"] 
            offset_local = item["offset"]
            offset_sum += offset_local
            break_points = [point if point <= items_row else point + offset_local for point in break_points]

        last_row = sheet.max_row + 2 # +2 строки для печати
        i = 1
        start_items_row = items_row - 2 # -2: Добавил шапку и 0 строку
        end_items_row = items_row + offset_sum + 1 # + 1: Добавил колонку итогов
        row_height_sum = 0
        start_id = 0
        max_height_on_list = 530 #530 - размер листа

        while i <= last_row:
            
            if not row_height_sum: #Для отладки
                start_id = i

            #Если превышаем максимальный размер, то разрываем на ближайшем break_point < i
            if sheet.row_dimensions[i].height:
                row_height_sum += sheet.row_dimensions[i].height
            else:
                row_height_sum += 11.25
            
            if row_height_sum < max_height_on_list:
                i += 1
            else:
                #print(f'Break:number: {i} - sum: {row_height_sum}')
                if start_items_row <= i <= end_items_row:
                    i += 1
                    row_height_sum = 0
                else:
                    break_point = self.find_nearest_lesser_or_equal(break_points, i)
                    if not break_point:
                        row_height_sum = 0
                        i += 1
                    else:
                        i = break_point
                        sheet.row_breaks.append(Break(id=i))

                        '''
                        #Отладочная код, возможно надо будет вернутся к функции
                        test_sum = 0
                        for j in range(start_id, i):
                            if sheet.row_dimensions[j].height:
                                test_sum += sheet.row_dimensions[j].height
                            else:
                                test_sum += 11.25
                        print(f'test sum = {test_sum}')
                        '''
                        i += 1
                        row_height_sum = 0
       #print(f'{start_id} : {i} : {row_height_sum}: {self.find_nearest_lesser_or_equal(break_points, i)}')

    def add_image(self, sheet: Worksheet, img_file_path: str, width: int, height: int, cell: str) -> None:
        """
        Вставляет картинку в таблицу excel
        """
        with default_storage.open(img_file_path, 'rb') as stamp_file:
                binary_image = BytesIO(stamp_file.read())
        img = XLImage(binary_image)
        img.width = width
        img.height = height
        sheet.add_image(img, cell)

    def float_to_kopecks(self, amount):
        # Получаем дробную часть числа
        fractional_part = amount % 1

        # Преобразуем её в целые копейки
        kopecks = round(fractional_part * 100)

        # Формирование строки с двумя цифрами копеек
        kopecks_str = f"{kopecks:02d}"

        # Склоняем слово "копейка"
        if 11 <= abs(kopecks) % 100 <= 14:
            word_form = 'копеек'
        elif abs(kopecks) % 10 == 1:
            word_form = 'копейка'
        elif 2 <= abs(kopecks) % 10 <= 4:
            word_form = 'копейки'
        else:
            word_form = 'копеек'

        return f'{kopecks_str} {word_form}'

    def add_document_itmes(self, sheet: Worksheet, items: dict, offsets: list, items_dict: dict, document: BaseModel = None,height_orientation_name: str = 'name', height_orientation_column: str = 'R') -> int:
        """
        Вставляет строки в таблицу excel и возвращает смещение
        """
        merge_areas = []
        merge_items = []
        merge_items_name = None
        offset = 0
        cell_itmes_number = items_dict["cell_itmes_number"]

        cell_itmes_number = self.get_row_with_offset(cell_itmes_number, offsets)

        #Заполняем список координат объедененных ячеек после строки с таблицы
        for mrg in sheet.merged_cells.ranges:
            start_col, start_row, end_col, end_row = range_boundaries(str(mrg))  # получаем границы объединённой области
            if start_row > cell_itmes_number: 
                merge_areas.append(mrg)
            elif start_row >= cell_itmes_number and end_row <= cell_itmes_number:
                merge_items.append(mrg)
                if start_col == column_index_from_string(height_orientation_column):
                    merge_items_name = mrg
                
        #Сохраняем высоту строк
        row_heights = {}
        for row in range(cell_itmes_number+1, sheet.max_row):
            row_heights[row] = sheet.row_dimensions[row].height

        for merge_range in merge_areas:
            sheet.unmerge_cells(str(merge_range))

        #Сохраняем стили ячеек
        styles_source = []
        for cell in sheet[f'A{cell_itmes_number}':f'{get_column_letter(sheet.max_column)}{cell_itmes_number}'][0]:
            styles_source.append({
                'font': cell.font.copy(),
                'fill': cell.fill.copy(),
                'border': cell.border.copy(),
                'alignment': cell.alignment.copy()
         })
        
        offset = len(items)-1#.count()
        sum = 0
        #Добавляем строки и заполняем
        for i, item in enumerate(items):
            number_of_row = cell_itmes_number + i
            if i:
                sheet.insert_rows(number_of_row)

            for col_idx, style in enumerate(styles_source):
                target_cell = sheet.cell(row=number_of_row, column=col_idx + 1)
                target_cell.font = style['font']
                target_cell.fill = style['fill']
                target_cell.border = style['border']
                target_cell.alignment = style['alignment']

            if "row_number" in items_dict:
                sheet[f'{items_dict["row_number"]}{number_of_row}'] = i + 1

            if "items_content" in items_dict:
                for key, cell_ref in items_dict["items_content"].items():
                    if key in item and item[key]:
                        value = str(item[key])
                        sheet[f'{cell_ref}{number_of_row}'] = value

                        if "sum_cell" in items_dict and cell_ref == items_dict["sum_cell"][0]:
                            sum += float(value)

                        if key == height_orientation_name:
                            
                            if merge_items_name:
                                coords = range_boundaries(str(merge_items_name))
                                column_width = int(coords[2]) - int(coords[0])
                                sheet.row_dimensions[cell_itmes_number + i].height = self.calculate_row_height(value, target_cell.font, column_width)

            if "static_content" in items_dict and document:
                for key, cell_ref in items_dict["static_content"].items():
                    value = self.get_nested_attribute(document, key)
                    sheet[f'{cell_ref}{number_of_row}'] = value

            for area in merge_items:
                coords = range_boundaries(str(area))
                adjusted_area = f'{get_column_letter(coords[0])}{number_of_row}:{get_column_letter(coords[2])}{number_of_row}'
                sheet.merge_cells(adjusted_area)

        if sum:
            sheet[f'{items_dict["sum_cell"][0]}{self.get_row_with_offset(items_dict["sum_cell"][1]+offset, offsets)}'] = f"{sum:.2f}"
            if "sum_str" in items_dict:
                print(int(sum))
                sheet[f'{items_dict["sum_str"][0]}{self.get_row_with_offset(items_dict["sum_str"][1]+offset, offsets)}'].value += rubles(int(sum), zero_for_kopeck=False).capitalize() + " " + self.float_to_kopecks(sum)

        #Возвращаем высоту строк после добавления
        for row, height in row_heights.items():
            sheet.row_dimensions[row + offset].height = height

        #Соединяем зоны после добавления строк
        for area in merge_areas:
            coords = range_boundaries(str(area))
            adjusted_area = f'{get_column_letter(coords[0])}{coords[1]+offset}:{get_column_letter(coords[2])}{coords[3]+offset}'
            sheet.merge_cells(adjusted_area)

        return offset
    
    def calculate_row_height(self, text: str, font: Font, column_width: float) -> float:
        """
        Расчитывает выстоту строки
        """
        line_spacing = 1.5

        lines = text.split('\n')
        #print(f"{column_width}, {font.size}")
        approx_chars_per_line = math.floor(column_width / (font.size / 7))

        if not approx_chars_per_line:
            #print(text)
            approx_chars_per_line = 1

        num_lines = 0

        for line in lines:
            num_lines += math.ceil(len(line) / approx_chars_per_line)

        #print(f'{approx_chars_per_line}:{num_lines}:{len(lines[0])}')
        return font.size  * num_lines * line_spacing
    
    def row_height_from_content(self, sheet: Worksheet, value: str, cell_ref: str) -> None:
        """
        Изменяет высоту строки на основе контента
        """
        col_number, row_number = coordinate_from_string(cell_ref)
        row_number = int(row_number)
        new_height = self.calculate_row_height(value, sheet[cell_ref].font, self.get_merged_column_count(sheet, cell_ref))
        if not sheet.row_dimensions[row_number].height or new_height > sheet.row_dimensions[row_number].height:
            sheet.row_dimensions[row_number].height = new_height
        #print(f'{cell_ref}: {value}: {sheet.row_dimensions[row_number].height}')

    def get_merged_column_count(self, sheet: Worksheet, cell_ref: str) -> int:
        """
        Функция возвращает количество столбцов у объедененной ячейки
        """
        col_number, row_number = coordinate_from_string(cell_ref)
        col_number = column_index_from_string(col_number)
        row_number = int(row_number)

        for mrg in sheet.merged_cells.ranges:
            start_col, start_row, end_col, end_row = range_boundaries(str(mrg))  # получаем границы объединённой области
            if start_col <= col_number <= end_col and start_row <= row_number <= end_row:
                return end_col - start_col
            
        return 1

    def read_and_return_file(self, file_path: str) -> BytesIO:
        with open(file_path, 'rb') as f:
            data = f.read()
        buffer = BytesIO(data)
        buffer.seek(0)
        return buffer

    def toPDF_libre(self, file_path: str) -> BytesIO:

        base_dir = os.path.dirname(file_path)
        filename = os.path.basename(file_path).split('.')[0]
        out_file = os.path.join(base_dir, f'{filename}.pdf')

        #libreoffice_path = 'C:\\Program Files\\LibreOffice\\program\\soffice.exe'
        libreoffice_path = '/usr/bin/soffice'

        if not os.path.exists(file_path):
            print(f"Ошибка! Файл '{file_path}' не найден.")
        else:
            command = [
                libreoffice_path, '--headless',
                '--convert-to', 'pdf',
                '--outdir', base_dir,
                file_path,
            ]
            
            try:
                result = subprocess.run(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
                if result.returncode != 0:
                    raise Exception(result.stderr.decode())
                else:
                    return self.read_and_return_file(out_file)
            except Exception as err:
                print(f"Ошибка при преобразовании: {err}")
